import re
import json
import os
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pandas as pd
import pickle
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime

#define a constant string
YES = 'Có'
NO = 'Không'
OWNER = 'Chủ hộ'

def string_to_timestamp(date_string, format='%m/%d/%Y %H:%M:%S'):
    """
    Convert a string to a timestamp.
    
    :param date_string: The date string to convert.
    :param format: The format of the date string.
    :return: A datetime object representing the timestamp.
    """
    date = datetime.strptime(date_string, format)
    return date.timestamp()


def normalize_full_name(text):
    """
    Capitalize the first letter of a string.
    """
    return ' '.join(word.capitalize() for word in text.split())


def normalize_date(text, inputFormat='%m/%d/%Y', outputFormat='%d/%m/%Y'):
    """
    Normalize a date string to the output format.
    """
    date = datetime.strptime(text, inputFormat)
    return date.strftime(outputFormat)


def excel_col_to_index(col_str):
    """
    Convert Excel column letters to 0-based column index.
    For example: A -> 0, B -> 1, Z -> 25, AA -> 26, AB -> 27, etc.
    """
    result = 0
    for char in col_str:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result - 1

def process_sheet_data(values, column_mapping):
    """
    Process sheet data into a pandas DataFrame with proper column mapping
    """
    if not values:
        return None

    # Get the maximum number of columns in any row
    max_cols = max(len(row) for row in values)
    
    # Get headers from first row, pad if necessary
    headers = values[0] if values else []
    if len(headers) < max_cols:
        headers.extend([f'Unnamed_{i+1}' for i in range(len(headers), max_cols)])
    
    # Pad all rows to have the same number of columns
    padded_values = []
    for idx, row in enumerate(values[1:], start=1):
        padded_row = row + [''] * (max_cols - len(row)) if len(row) < max_cols else row
        # append original row index into the last column
        padded_row.append(idx)
        padded_values.append(padded_row)

    # sort the rows by the block, floor, home accending and timestamp descending
    blockCol = excel_col_to_index('B')
    floorCol = excel_col_to_index('C')
    homeCol = excel_col_to_index('D')
    timestampCol = excel_col_to_index('A')
    padded_values.sort(key=lambda x: (createHomeID(x[blockCol], x[floorCol], x[homeCol]), -(string_to_timestamp(x[timestampCol]))), reverse=False)

    # warning if dupplicate block, floor, home
    dupplicateRows = []
    originalRow = -1
    for i in range(1, len(padded_values)):
        block = padded_values[i][blockCol]
        floor = padded_values[i][floorCol]
        home = padded_values[i][homeCol]
        row = padded_values[i][-1]
        if block == padded_values[i-1][blockCol] and floor == padded_values[i-1][floorCol] and home == padded_values[i-1][homeCol]:
            # set original row
            if  originalRow == -1:
                originalRow = padded_values[i-1][-1]
            print(f"* Warning: Dupplicate {createHomeID(block, floor, home)} at row {row + 1} vs lastest {originalRow + 1}")
            dupplicateRows.append(i)
        else:
            # reset original row
            originalRow = -1

    # remove dupplicate rows
    for i in range(len(dupplicateRows)-1, -1, -1):
        padded_values.pop(dupplicateRows[i])

    # Process each row into a DataFrame
    split_dfs = []
    additional_info_data = [''] * len(column_mapping['dest_additional_info_ids'])

    for idx, row in enumerate(padded_values, start=0):
        try:
            # Extract common info using the correct column indices
            common_info = []
            for col in column_mapping['src_common_info_id']:
                col_idx = excel_col_to_index(col)
                common_info.append(row[col_idx] if col_idx < len(row) else '')

            # Merge floor and home into a single column
            block = common_info[1]
            floor = common_info[2]
            home = common_info[3]
            common_info[2] = createHomeID(block, floor, home)
            common_info.pop(3)
            
            # Replace the first element with the row index
            common_info[0] = idx + 1
            
            # Extract owner info using the correct column indices
            owner_info = []
            for col in column_mapping['src_owner_info_id']:
                col_idx = excel_col_to_index(col)
                owner_info.append(row[col_idx] if col_idx < len(row) else '')
            # Append the last element with 'Owner'
            owner_info.append(OWNER)
            # capitalize the first letter of the full name
            owner_info[0] = normalize_full_name(owner_info[0])
            # normalize the date format
            owner_info[2] = normalize_date(owner_info[2])
            

            # G-Row ID
            additional_info_data[-1] = row[-1] + 1

            # Create DataFrame for the row
            row_data = [common_info + owner_info + additional_info_data]
            row_columns = column_mapping['dest_common_info_names'] + column_mapping['dest_member_info_names'] + column_mapping['dest_additional_info_names']
            row_df = pd.DataFrame(row_data, columns=row_columns)
            split_dfs.append(row_df)

            # Extract member info using the correct column indices
            hasMember = row[excel_col_to_index(column_mapping['src_owner_info_next_id'])] == YES
            if hasMember:
                member_mapping = {
                    'src_member_info_next_id': column_mapping['src_member_info_next_id'],
                    'src_member_info_names': column_mapping['src_member_info_names'],
                    'src_member_info_id': column_mapping['src_member_info_id'],
                    'dest_member_info_ids': column_mapping['dest_member_info_ids'],
                    'dest_member_info_names': column_mapping['dest_member_info_names']
                }
                members_info = process_member_data(row, member_mapping)

                for member_info in members_info:
                    row_member_data = [common_info + member_info + additional_info_data]
                    row_member_columns = column_mapping['dest_common_info_names'] + column_mapping['dest_member_info_names'] + column_mapping['dest_additional_info_names']
                    row_member_df = pd.DataFrame(row_member_data, columns=row_member_columns)
                    split_dfs.append(row_member_df)
            
        except Exception as e:
            print(f"* Error processing row {idx}: {str(e)}")
            continue

    if not split_dfs:
        return None

    return pd.concat(split_dfs, ignore_index=True)


def regex_extract_number(text):
    """
    Extract the number from a string using regex and pad it with leading zeros to length 2.
    """
    match = re.search(r'\d+', text)
    if match:
        number = match.group(0)
        return number.zfill(2)  # Pad the number with leading zeros to ensure length is 2
    return '00'  # Return '00' if no number is found


def createHomeID(block, floor, home):
    """
    Create a HomeID from the block, floor, and home
    block = C1
    floor = Tầng 12
    home = Căn hộ 07
    HomeID = C1-1207
    """

    # return f'{block}-{floor}-{home}'

    # floorNumber = floor.split(' ')[1]
    # homeNumber = home.split(' ')[2]
    # return f'{block}-{floorNumber}{homeNumber}'

    return f'{block}-{regex_extract_number(floor)}{regex_extract_number(home)}'


def process_member_data(row, member_mapping):
    """
    Process member data into a list of rows for the member DataFrame
    """
    member_info = []
    member_info_next_id = member_mapping['src_member_info_next_id']
    member_info_names = member_mapping['src_member_info_names']
    member_info_ids = member_mapping['src_member_info_id']
    member_info_next_col = excel_col_to_index(member_info_next_id)
    member_info_columns = member_mapping['dest_member_info_names']
    
    col_offset = 0

    while True:
        member_data = []
        
        # Extract member info using the correct column indices
        for col in member_info_ids:
            col_idx = excel_col_to_index(col) + col_offset
            member_data.append(row[col_idx] if col_idx < len(row) else '')
        # swap the last two elements
        # because the last element is the phone number
        member_data[-1], member_data[-2] = member_data[-2], member_data[-1]
        # capitalize the first letter of the full name
        member_data[0] = normalize_full_name(member_data[0])
        # normalize the date format
        member_data[2] = normalize_date(member_data[2])
        
        # Append the member info to the list
        member_info.append(member_data)
        
        # Check if there are more members
        if (member_info_next_col + col_offset) >= len(row) or row[member_info_next_col + col_offset] == NO:
            break

        # Move to the next member
        col_offset  += len(member_info_ids) + 1

    return member_info


def get_credentials():
    """
    Get cached credentials or create new ones if needed.
    """
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = None
    token_path = 'token.pickle'
    
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)
    
    return creds


def merge_cells_if_same(sheet, start_row, end_row, col):
    """
    Merge cells in the specified column if they have the same value.
    """
    merge_start = start_row
    for row in range(start_row, end_row + 1):
        if row == end_row or sheet.cell(row=row, column=col).value != sheet.cell(row=row + 1, column=col).value:
            if merge_start != row:
                sheet.merge_cells(start_row=merge_start, start_column=col, end_row=row, end_column=col)
            merge_start = row + 1


def post_process_and_save_to_excel(output_path, merge_col_names, group_by_col):
    """
    Process Excel file and merging cells with the same value.
    """
    # Load the workbook and the first sheet
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    max_row = ws.max_row + 1
    group_by_col_index = excel_col_to_index(group_by_col) + 1

    fill_colors = ['DDFFFF', 'FFFFFF']  # Light blue and white
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row = 2
    while row < max_row:
        group_by_value = ws.cell(row=row, column=group_by_col_index).value

        merge_start = row
        while row < max_row and ws.cell(row=row, column=group_by_col_index).value == group_by_value:
            row += 1
        merge_end = row - 1

        # Apply fill color to the group
        stt_value = ws.cell(row=merge_start, column=1).value
        color_index = int(stt_value) % len(fill_colors)
        fill = PatternFill(start_color=fill_colors[color_index], end_color=fill_colors[color_index], fill_type="solid")
        for r in range(merge_start, merge_end + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill
                ws.cell(row=r, column=c).border = thin_border

        if merge_start < merge_end:
            # Merge cells in column if they have the same value
            # print(f"* Merging cells for group '{group_by_value}' from rows {merge_start} to {merge_end}")
            for col_name in merge_col_names:
                col_index = excel_col_to_index(col_name) + 1
                merge_cells_if_same(ws, merge_start, merge_end, col_index)

    # Save the workbook
    wb.save(output_path)


def gsheet_to_xlsx(gsheet_path, output_path, column_mapping):
    """
    Convert a .gsheet file to .xlsx format, handling column mismatches
    """
    try:
        # Read the .gsheet file
        with open(gsheet_path, 'r') as f:
            gsheet_data = json.load(f)
        
        spreadsheet_id = gsheet_data.get('doc_id')
        if not spreadsheet_id:
            raise ValueError("Could not find spreadsheet ID in .gsheet file")
        
        # Get cached credentials
        creds = get_credentials()
        
        # Build the Sheets API service
        service = build('sheets', 'v4', credentials=creds)
        
        # Get all sheets in the spreadsheet
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        
        if not sheets:
            raise ValueError("No sheets found in the spreadsheet")
        
        # Filter for visible sheets
        visible_sheets = [sheet for sheet in sheets 
                        if not sheet.get('properties', {}).get('hidden', False)]
        
        if not visible_sheets:
            print("* Warning: All sheets are hidden. Attempting to use all sheets instead.")
            visible_sheets = sheets
        
        print(f"* Found {len(visible_sheets)} sheet(s)")
        
        # Create Excel writer
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        sheets_processed = 0
        
        try:
            for sheet in visible_sheets:
                properties = sheet['properties']
                sheet_name = properties['title']
                print(f"* Processing sheet: {sheet_name}")
                
                try:
                    # Get the sheet data
                    result = service.spreadsheets().values().get(
                        spreadsheetId=spreadsheet_id,
                        range=f'{sheet_name}!A1:ZZ'
                    ).execute()
                    
                    values = result.get('values', [])
                    if not values:
                        print(f"* Sheet '{sheet_name}' is empty, skipping")
                        continue
                    
                    # Process the sheet data
                    final_df = process_sheet_data(values, column_mapping)
                    if final_df is None:
                        print(f"* No valid data in sheet '{sheet_name}', skipping")
                        continue
                    
                    # Write to Excel
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    sheets_processed += 1
                    print(f"* Successfully processed sheet: {sheet_name}")
                    
                except Exception as e:
                    print(f"* Error processing sheet '{sheet_name}': {str(e)}")
                    continue
            
            if sheets_processed == 0:
                raise ValueError("No sheets could be processed successfully")
                       
            print(f"* Successfully converted {gsheet_path} to {output_path}")
            print(f"* Total sheets processed: {sheets_processed}")
            
        finally:
            # Make sure we close the writer even if an error occurs
            if writer:
                writer.close()

        post_process_and_save_to_excel(output_path, column_mapping['dest_merge_cells_ids'], column_mapping['dest_group_by_id']) 
                
    except Exception as e:
        print(f"* An error occurred: {str(e)}")
        raise

# Example usage showing how to handle columns beyond Z
if __name__ == "__main__":
    column_mapping = {
        'src_common_info_id': ['A', 'B', 'C', 'D', 'E'],
        'src_common_info_names': ['Timestamp', 'Block', 'Floor', 'Home', 'Owner'],
        
        'src_owner_info_id': ['F', 'G', 'H', 'I'],
        'src_owner_info_names': ['Full Name', 'Sex', 'Birthday', 'Phone'],
        
        'src_owner_info_next_id': 'J',

        'src_member_info_id': ['K', 'L', 'M', 'N', 'O'],
        'src_member_info_names': ['Full Name', 'Sex', 'Birthday', 'Relationship', 'Phone'],
        'src_member_info_next_id': 'P',
        
        'dest_common_info_ids': ['A', 'B', 'C', 'D'],
        'dest_common_info_names': ['STT', 'BLOCK', 'MÃ CĂN HỘ', 'CHÍNH CHỦ/THUÊ'],
        
        'dest_member_info_ids': ['E', 'F', 'G', 'H', 'I'],
        'dest_member_info_names': ['HỌ VÀ TÊN', 'GIỚI TÍNH', 'NGÀY/THÁNG/NĂM SINH', 'SĐT', 'QH VỚI CHỦ HỘ/NGƯỜI THUÊ'],

        'dest_additional_info_ids': ['J', 'K', 'L'],
        'dest_additional_info_names': ['THÔNG TIN CHỦ CŨ', 'THÔNG TIN CHỦ HỘ', 'G-Row ID'],

        'dest_merge_cells_ids': ['A', 'B', 'C', 'D', 'J', 'K', 'L'],
        'dest_merge_cells_names': ['STT', 'BLOCK', 'MÃ CĂN HỘ', 'CHÍNH CHỦ/THUÊ', 'THÔNG TIN CHỦ CŨ', 'THÔNG TIN CHỦ HỘ', 'G-Row ID'],

        'dest_group_by_id': 'C',
        'dest_group_by_name': 'MÃ CĂN HỘ'
    }

    gsheet_to_xlsx("input.gsheet", "output.xlsx", column_mapping)

